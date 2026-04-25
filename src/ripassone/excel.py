"""Parser del template Excel "DIPA L19 - compito".

Il foglio "Domande" ha header alla riga 7:
  # | Lezione | Argomento | Domanda | Opzione A | B | C | D
  | Risposta corretta | Difficolta | Spiegazione | Fonte

Le righe utili partono dalla 8. La cella B3 contiene "Cognome e nome"
dello studente (autore delle domande).

Il parser tollera valori parziali e righe vuote: ritorna le domande
valide + una lista di errori per le righe scartate (con riferimento
alla riga originale, utile per il feedback all'admin).
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import IO

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from ripassone.models import Question, QuestionOption

SHEET_NAME = "Domande"
HEADER_ROW = 7
DATA_START_ROW = 8
AUTHOR_CELL = "B3"


@dataclass
class ParseError:
    row: int
    field: str | None
    message: str

    def to_dict(self) -> dict:
        return {"row": self.row, "field": self.field, "message": self.message}


@dataclass
class ParseResult:
    author: str
    questions: list[Question]
    errors: list[ParseError]
    total_rows_scanned: int

    def to_dict(self) -> dict:
        return {
            "author": self.author,
            "added": len(self.questions),
            "errors": [e.to_dict() for e in self.errors],
            "total_rows_scanned": self.total_rows_scanned,
        }


def parse_workbook(content: bytes | IO[bytes], filename: str = "", id_offset: int = 0) -> ParseResult:
    """Parsifica un file Excel. id_offset e usato dal chiamante per assegnare
    id univoci continuativi (max(pool_ids) + 1)."""
    if isinstance(content, bytes):
        bio = io.BytesIO(content)
    else:
        bio = content
    try:
        wb: Workbook = load_workbook(bio, data_only=True, read_only=True)
    except Exception as e:
        return ParseResult(
            author="?", questions=[], total_rows_scanned=0,
            errors=[ParseError(row=0, field=None, message=f"File non leggibile come Excel: {e}")],
        )

    if SHEET_NAME not in wb.sheetnames:
        return ParseResult(
            author="?", questions=[], total_rows_scanned=0,
            errors=[ParseError(row=0, field=None,
                               message=f'Foglio "{SHEET_NAME}" non trovato (presenti: {wb.sheetnames})')],
        )

    ws = wb[SHEET_NAME]

    # autore: B3 oppure dal filename oppure "anonimo"
    author = _coerce_str(ws[AUTHOR_CELL].value) or _author_from_filename(filename) or "anonimo"

    questions: list[Question] = []
    errors: list[ParseError] = []
    next_id = id_offset + 1
    rows_scanned = 0

    # iter dalle righe dati in poi
    # read_only mode: usiamo iter_rows per efficienza
    for raw_row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        rows_scanned += 1
        excel_row = DATA_START_ROW + rows_scanned - 1

        if not any(cell is not None and str(cell).strip() != "" for cell in raw_row):
            continue  # riga totalmente vuota, skip silenzioso

        # padding a 12 colonne se per qualche motivo ne arrivano meno
        cells = list(raw_row) + [None] * (12 - len(raw_row))
        _, lecture, topic, text, oa, ob, oc, od, correct, diff, _spieg, source = cells[:12]

        text_s = _coerce_str(text)
        # se solo la colonna # e' compilata (template "vuoto"), skip silenzioso
        other_filled = any(_coerce_str(c) for c in (lecture, topic, oa, ob, oc, od, correct, diff))
        if not text_s and not other_filled:
            continue
        if not text_s:
            errors.append(ParseError(row=excel_row, field="Domanda", message="testo domanda vuoto"))
            continue

        opts_text = [_coerce_str(o) for o in (oa, ob, oc, od)]
        if not all(opts_text):
            missing = [letter for letter, ot in zip("ABCD", opts_text) if not ot]
            errors.append(ParseError(
                row=excel_row, field="Opzione " + "/".join(missing),
                message=f"opzioni mancanti: {missing}",
            ))
            continue

        correct_letter = _parse_correct_letter(correct, opts_text)
        if correct_letter is None:
            errors.append(ParseError(
                row=excel_row, field="Risposta corretta",
                message=f"valore non valido: {correct!r} (atteso A/B/C/D o testo opzione)",
            ))
            continue

        difficulty = _parse_difficulty(diff)

        lecture_s = _coerce_str(lecture) or "?"
        topic_s = _coerce_str(topic) or ""
        source_s = _coerce_str(source)

        q = Question(
            id=next_id,
            lecture=lecture_s,
            topic=topic_s,
            text=text_s,
            options=[QuestionOption(letter=l, text=t) for l, t in zip("ABCD", opts_text)],
            correct=correct_letter,  # type: ignore[arg-type]
            difficulty=difficulty,
            source=source_s,
            author=author,
        )
        questions.append(q)
        next_id += 1

    wb.close()
    return ParseResult(author=author, questions=questions, errors=errors, total_rows_scanned=rows_scanned)


# ============================================================
# Helpers di parsing tolleranti
# ============================================================
def _coerce_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _author_from_filename(filename: str) -> str:
    """Estrae nome autore dal filename (es. 'compito_RossiMario.xlsx' -> 'RossiMario')."""
    if not filename:
        return ""
    name = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0]
    # qualche tentativo di splittare convenzioni note
    for prefix in ("DIPA_L19_compito_", "compito_", "L19_", "DIPA_"):
        if name.startswith(prefix):
            return name[len(prefix):]
    return name


def _parse_correct_letter(value, opts_text: list[str]) -> str | None:
    """Riconosce A/B/C/D. Se invece riceve il testo dell'opzione, fa il match."""
    s = _coerce_str(value).upper()
    if s in ("A", "B", "C", "D"):
        return s
    # match per testo (case-insensitive)
    s_lower = s.lower()
    for letter, text in zip("ABCD", opts_text):
        if text.lower() == s_lower:
            return letter
    return None


def _parse_difficulty(value) -> int:
    """Accetta 1/2/3 o 'Facile'/'Media'/'Tosta' (case-insensitive). Default 2."""
    if value is None:
        return 2
    s = _coerce_str(value).lower()
    if s in ("1", "facile", "f", "easy", "low"):
        return 1
    if s in ("2", "media", "medium", "m", "mid"):
        return 2
    if s in ("3", "tosta", "difficile", "hard", "h", "tough"):
        return 3
    # tentativo numerico
    try:
        n = int(float(s))
        if 1 <= n <= 3:
            return n
    except ValueError:
        pass
    return 2
